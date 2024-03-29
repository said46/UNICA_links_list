from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import InvalidSelectorException
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service
from message import message_box
import os
import time
import openpyxl as xl

def get_links_from_page(msg: str=""):
    # message_box('Message', 'Retrieving links from ' + msg, 0)
    element_xpath = f"//a[@class='browseItemNameContainer']"
    try:
        elem_list = edgeBrowser.find_elements(By.XPATH, element_xpath)
        # print(f'Processing each element...') 
        for e in elem_list:
            doc_list.append({"loop_name": e.get_attribute(name='innerText')[:-4], "loop_link": e.get_attribute(name='href')})        
    except NoSuchElementException:
        message_box('Error', 'NoSuchElementException', 0)        
    except StaleElementReferenceException:
        message_box('Error', 'StaleElementReferenceException', 0)
        edgeBrowser.quit()
        quit()


"""
    The scipt opens the UNICA link (folder) and saves all the links (even if there are multiple pages for 125+ files)
    It it also required to specify the number of files (links_count variable)
"""

# page_link = f'https://sww-llsak.sakhalinenergy.ru/glasseic/livelink.exe/Open/140110499'
page_link = f'https://sww-llsak.sakhalinenergy.ru/glasseic/livelink.exe/Open/140135351'
links_count = 736
excel_columns = {"loop_name": 1, "loop_link": 2}
excel_filename = 'links.xlsx'

os.system("cls")
script_dir = os.path.abspath(os.path.dirname( __file__ ))

print(f'Opening the Excel-file...')
wb_fullpath = script_dir + "\\" + excel_filename
try:
    wb = xl.load_workbook(wb_fullpath)
except Exception as e:
    print(f'Cannot open the excel file: {str(e)}, quitting...')
    message_box('Error', f'Cannot open the excel file: {str(e)}', 0)    
    quit()
print(f'{excel_filename} has been opened successfully')

ws = wb['links']

# clear prev. data
print(f'Clearing data in {excel_filename}...')
for row in ws['A2':'B5000']:
  for cell in row:
    cell.value = None
    cell.hyperlink = None

print(f'Starting Edge...')
options = EdgeOptions()
options.add_argument("start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
service_path = script_dir + "\\msedgedriver.exe"
edgeBrowser = webdriver.Edge(service=Service(service_path), options=options)
edgeBrowser.implicitly_wait(5)

doc_list = list()

# open the starting page
try:
    edgeBrowser.get(page_link)
except Exception as e:
    print(f"{str(e)}, aborting the program")
    edgeBrowser.quit()
    quit()

print(f'Processing starting page...')
get_links_from_page("starting page")

for page in range(2, (links_count + 201) // 100):
    element_xpath = f'//td[@title="Page {page}"]'
    try:
        elem = edgeBrowser.find_element(By.XPATH, element_xpath)
        # elem = WebDriverWait(edgeBrowser, 5).until(ec.element_to_be_clickable((By.XPATH, element_xpath)))
        elem.click()
    except NoSuchElementException:
        message_box('Error', 'NoSuchElementException', 0)
    time.sleep(3)
    print(f'Processing page {page}...')
    get_links_from_page(f"page {page}")
        
row = 2
for doc in doc_list:   
    ws.cell(row, excel_columns["loop_name"]).value = doc["loop_name"]
    ws.cell(row, excel_columns["loop_name"]).hyperlink = doc["loop_link"]
    ws.cell(row, excel_columns["loop_link"]).value = doc["loop_link"]
    row += 1
    
print(f'Saving the Excel-file...')
try:
  wb.save(wb_fullpath)
except Exception as e:
    print(f'Cannot save the excel file: {str(e)}')

wb.close()
print(f"End of script")
message_box('End of script', 'Ok!', 0)
edgeBrowser.quit()